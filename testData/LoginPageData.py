import openpyxl


class LoginPageData:
    test_LoginPage_Data = [{"email": "anantbinekar@gmail.com", "password": "any@0000"}]

@staticmethod
def get_test_data(test_case_name):
    Dict = {}
    book = openpyxl.load_workbook("C:\\Users\\HP\\PycharmProjects\\PythonSeleniumFramework\\testData\\PythonDemo.xlsx")
    sheet = book.active
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

    return [Dict]
